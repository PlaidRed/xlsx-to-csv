from flask import Flask, request, jsonify, send_file
from werkzeug.utils import secure_filename
import openpyxl
import csv
import os
import uuid
import zipfile
import traceback
from pathlib import Path

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size
app.config['UPLOAD_FOLDER'] = 'uploads'

# Disable Flask's default HTML error pages
app.config['PROPAGATE_EXCEPTIONS'] = True

# Create uploads directory if it doesn't exist
Path(app.config['UPLOAD_FOLDER']).mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return send_file('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        # Check if file is present
        if 'excel_file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400
        
        file = request.files['excel_file']
        
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'message': 'Invalid file type. Only .xlsx and .xls allowed'}), 400
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4()}-{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)
        
        # Get original file size
        original_size = round(os.path.getsize(filepath) / 1024, 2)
        
        # Load Excel file (read-only mode for better performance)
        workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet = workbook.active
        
        # Get dimensions
        max_row = sheet.max_row
        max_column = sheet.max_column
        max_column_letter = openpyxl.utils.get_column_letter(max_column)
        
        # Read all data
        data = []
        for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column, values_only=True):
            data.append(row)
        
        workbook.close()
        
        # Find columns with any data
        columns_with_data = set()
        for row in data:
            for col_idx, cell in enumerate(row):
                if cell is not None and str(cell).strip() != '':
                    columns_with_data.add(col_idx)
        
        active_columns = sorted(list(columns_with_data))
        
        # Create CSV file
        csv_filename = f"{uuid.uuid4()}.csv"
        csv_filepath = os.path.join(app.config['UPLOAD_FOLDER'], csv_filename)
        
        row_count = 0
        empty_rows_skipped = 0
        
        with open(csv_filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            for row in data:
                # Check if row has any content in active columns
                has_content = False
                for col_idx in active_columns:
                    if col_idx < len(row):
                        value = str(row[col_idx]).strip() if row[col_idx] is not None else ''
                        if value != '':
                            has_content = True
                            break
                
                # Skip empty rows
                if not has_content:
                    empty_rows_skipped += 1
                    continue
                
                # Extract only active columns
                cleaned_row = []
                for col_idx in active_columns:
                    if col_idx < len(row):
                        value = str(row[col_idx]).strip() if row[col_idx] is not None else ''
                        cleaned_row.append(value)
                    else:
                        cleaned_row.append('')
                
                writer.writerow(cleaned_row)
                row_count += 1
        
        # Get CSV file size
        csv_size = round(os.path.getsize(csv_filepath) / 1024, 2)
        
        # Create ZIP file
        zip_filename = f"{uuid.uuid4()}.zip"
        zip_filepath = os.path.join(app.config['UPLOAD_FOLDER'], zip_filename)
        
        with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED, compresslevel=9) as zipf:
            # Add CSV to zip with original base name (without UUID)
            original_base = os.path.splitext(filename)[0]
            zipf.write(csv_filepath, f"{original_base}.csv")
        
        # Get ZIP file size
        zip_size = round(os.path.getsize(zip_filepath) / 1024, 2)
        
        # Calculate compression ratio
        compression_ratio = round((zip_size / csv_size) * 100, 1) if csv_size > 0 else 0
        
        # Clean up temporary files
        try:
            os.remove(filepath)  # Remove original Excel
            os.remove(csv_filepath)  # Remove uncompressed CSV
        except:
            pass
        
        return jsonify({
            'success': True,
            'original_size': original_size,
            'csv_size': csv_size,
            'zip_size': zip_size,
            'compression_ratio': compression_ratio,
            'zip_file': f'uploads/{zip_filename}',
            'rows_written': row_count,
            'excel_dimensions': f'{max_column_letter}{max_row}',
            'total_columns_in_excel': max_column,
            'columns_with_data': len(active_columns),
            'empty_rows_skipped': empty_rows_skipped
        })
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/uploads/<filename>')
def download_file(filename):
    try:
        return send_file(
            os.path.join(app.config['UPLOAD_FOLDER'], filename),
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'success': False, 'message': f'File not found: {str(e)}'}), 404

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)